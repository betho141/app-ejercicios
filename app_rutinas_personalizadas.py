
import streamlit as st
import pandas as pd
import webbrowser
import sqlite3
from openpyxl import load_workbook

# --- CONFIGURACIÓN ---
st.set_page_config(page_title="App de Ejercicios", layout="centered")

# --- CONEXIÓN A BASE DE DATOS ---
def obtener_id_usuario(nombre_clave):
    conn = sqlite3.connect("rutinas_personalizadas.db")
    cur = conn.cursor()
    cur.execute("SELECT id_usuario FROM usuarios WHERE nombre_clave = ?", (nombre_clave,))
    result = cur.fetchone()
    conn.close()
    return result[0] if result else None

def obtener_ids_ejercicios_por_usuario(id_usuario):
    conn = sqlite3.connect("rutinas_personalizadas.db")
    cur = conn.cursor()
    cur.execute("SELECT id_ejercicio FROM rutinas WHERE id_usuario = ?", (id_usuario,))
    resultados = [r[0] for r in cur.fetchall()]
    conn.close()
    return resultados

# --- LOGO LATERAL IZQUIERDO ---
def agregar_logo_lateral(imagen_url):
    st.markdown(
        f"""
        <style>
        .stApp {{
            background-image: url("{imagen_url}");
            background-repeat: no-repeat;
            background-attachment: fixed;
            background-size: 300px;
            background-position: 20px 40px;
            padding-left: 340px;
        }}
        </style>
        """,
        unsafe_allow_html=True
    )

# --- LOGO LATERAL DERECHO ---
def agregar_logo_derecho(imagen_url):
    st.markdown(
        f"""
        <style>
        .stApp::after {{
            content: "";
            background: url("{imagen_url}") no-repeat;
            background-size: 300px;
            position: fixed;
            top: 40px;
            right: 20px;
            width: 300px;
            height: 300px;
            z-index: 1;
        }}
        </style>
        """,
        unsafe_allow_html=True
    )

agregar_logo_lateral("https://i.imgur.com/N2ZpK7M.png")
agregar_logo_derecho("https://i.imgur.com/N2ZpK7M.png")  # Cambiar por otra si lo deseas

# --- AUTENTICACIÓN ---
st.title("🏋️‍♂️ App de Ejercicios Personalizados")

nombre_clave = st.text_input("🔐 Ingresa tu nombre clave:")

if nombre_clave:
    id_usuario = obtener_id_usuario(nombre_clave)

    if id_usuario:
        st.success(f"Bienvenido, {nombre_clave} 💪")
        ids_permitidos = obtener_ids_ejercicios_por_usuario(id_usuario)

        # --- CARGA DE EJERCICIOS DESDE EXCEL ---
        @st.cache_data
        def cargar_ejercicios(path):
            wb = load_workbook(path, data_only=True)
            ws = wb.active

            ejercicios = []
            for idx, row in enumerate(ws.iter_rows(min_row=2), start=1):
                cell = row[0]
                zona = row[2].value if len(row) >= 3 else ""
                implemento = row[4].value if len(row) >= 5 else ""
                articularidad = row[5].value if len(row) >= 6 else ""
                nombre = cell.value
                url = cell.hyperlink.target if cell.hyperlink else None
                if nombre and url:
                    ejercicios.append({
                        "id_ejercicio": idx,
                        "nombre": nombre,
                        "url": url,
                        "zona_corporal": zona,
                        "implemento": implemento,
                        "articularidad": articularidad
                    })
            return pd.DataFrame(ejercicios)

        df = cargar_ejercicios("All exercices.xlsx")

        # --- FILTRAR EJERCICIOS POR USUARIO ---
        df = df[df["id_ejercicio"].isin(ids_permitidos)]

        # --- BÚSQUEDA ---
        busqueda = st.text_input("🔍 Buscar ejercicio por nombre")
        if busqueda:
            df = df[df["nombre"].str.contains(busqueda, case=False)]

        # --- MOSTRAR TABLA ---
        st.subheader("📊 Tus Ejercicios Asignados")
        st.dataframe(df[["id_ejercicio", "nombre", "zona_corporal", "implemento", "articularidad"]]
                     .rename(columns={
                         "nombre": "Ejercicio",
                         "zona_corporal": "Zona Corporal",
                         "implemento": "Implemento",
                         "articularidad": "Articularidad"
                     }), use_container_width=True)

        # --- BOTONES DE VIDEO ---
        st.subheader("🎥 Ver Videos")
        for _, row in df.iterrows():
            if st.button(row["nombre"]):
                webbrowser.open_new_tab(row["url"])
    else:
        st.error("Nombre clave incorrecto. Por favor verifica tu acceso.")
