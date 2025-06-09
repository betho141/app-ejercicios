
import streamlit as st
import pandas as pd
import webbrowser
from openpyxl import load_workbook

# --- CONFIGURACIÓN ---
st.set_page_config(page_title="App de Ejercicios", layout="centered")

# --- LOGO ---
# st.image("logo.jpg", width=180)

# --- FONDO PERSONALIZADO ---
# def agregar_fondo(imagen_url):
#     st.markdown(
#         f"""
#         <style>
#         .stApp {{
#             background-image: url("{imagen_url}");
#             background-size: cover;
#             background-repeat: no-repeat;
#             background-position: center;
#         }}
#         </style>
#         """,
#         unsafe_allow_html=True
#     )

# agregar_fondo("https://i.imgur.com/4HJbzEq.jpeg")

def agregar_logo_lateral(imagen_url):
    st.markdown(
        f"""
        <style>
        .stApp {{
            background-image: url("{imagen_url}");
            background-repeat: no-repeat;
            background-attachment: fixed;
            background-size: 220px;
            background-position: top left;
            padding-left: 240px;  /* desplaza el contenido a la derecha */
        }}
        </style>
        """,
        unsafe_allow_html=True
    )
agregar_logo_lateral("https://imgur.com/a/fVlzXx2")  # o usa el PNG que subiste

# --- TÍTULO ---
st.title("🏋️‍♂️ App de Ejercicios con Videos")

# --- CARGA DE DATOS DESDE EXCEL ---
@st.cache_data
def cargar_datos_excel(path):
    from openpyxl import load_workbook
    wb = load_workbook(path, data_only=True)
    ws = wb.active

    ejercicios = []
    for row in ws.iter_rows(min_row=2):
        cell = row[0]
        zona = row[2].value if len(row) >= 3 else ""
        implemento = row[4].value if len(row) >= 5 else ""
        articularidad = row[5].value if len(row) >= 6 else ""
        nombre = cell.value
        url = cell.hyperlink.target if cell.hyperlink else None
        if nombre and url:
            ejercicios.append({
                "nombre": nombre,
                "url": url,
                "zona_corporal": zona,
                "implemento": implemento,
                "articularidad": articularidad
            })
    return pd.DataFrame(ejercicios)

df = cargar_datos_excel("All exercices.xlsx")

# --- FILTROS MULTIPLES ---
st.markdown("### Filtros para encontrar tu ejercicio ideal")
col1, col2, col3 = st.columns(3)

with col1:
    zonas_disponibles = ["Todas"] + sorted(df["zona_corporal"].dropna().unique().tolist())
    zona_seleccionada = st.selectbox("🧍‍♀️ Zona corporal", zonas_disponibles)

with col2:
    implementos_disponibles = ["Todos"] + sorted(df["implemento"].dropna().unique().tolist())
    implemento_seleccionado = st.selectbox("🏋️‍♂️ Implemento", implementos_disponibles)

with col3:
    articularidades_disponibles = ["Todas"] + sorted(df["articularidad"].dropna().unique().tolist())
    articularidad_seleccionada = st.selectbox("🔗 Articularidad", articularidades_disponibles)

# Aplicar filtros
if zona_seleccionada != "Todas":
    df = df[df["zona_corporal"] == zona_seleccionada]
if implemento_seleccionado != "Todos":
    df = df[df["implemento"] == implemento_seleccionado]
if articularidad_seleccionada != "Todas":
    df = df[df["articularidad"] == articularidad_seleccionada]

# --- BÚSQUEDA POR TEXTO ---
busqueda = st.text_input("🔍 Buscar ejercicio por nombre")

if busqueda:
    df = df[df["nombre"].str.contains(busqueda, case=False)]

# --- TABLA DE RESULTADOS ---
st.subheader("📊 Tabla de Ejercicios Filtrados")
st.dataframe(df[["nombre", "zona_corporal", "implemento", "articularidad"]]
             .rename(columns={
                 "nombre": "Ejercicio",
                 "zona_corporal": "Zona Corporal",
                 "implemento": "Implemento",
                 "articularidad": "Articularidad"
             }), use_container_width=True)

# --- BOTONES DE VIDEO ---
st.subheader("🎥 Ver Videos")
for _, row in df.iterrows():
    if st.button(row["nombre"]):
        webbrowser.open_new_tab(row["url"])
