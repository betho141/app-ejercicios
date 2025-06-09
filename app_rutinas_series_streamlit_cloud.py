import streamlit as st
import shutil
import os

DB_ORIG = "rutinas_personalizadas.db"
DB_PATH = "/tmp/rutinas_personalizadas.db"

# Copiar base de datos si no existe aún en /tmp
if not os.path.exists(DB_PATH):
    shutil.copy(DB_ORIG, DB_PATH)

import sqlite3
import pandas as pd
import webbrowser
from openpyxl import load_workbook

# --- CONFIGURACION ---
st.set_page_config(page_title="App Rutinas Personalizadas", layout="wide")

# --- FUNCION PARA CARGAR EXCEL ---
@st.cache_data
def cargar_datos_excel(path):
    wb = load_workbook(path, data_only=True)
    ws = wb.active
    ejercicios = []
    for row in ws.iter_rows(min_row=2):
        cell = row[0]
        zona = row[2].value if len(row) >= 3 else ""
        implemento = row[4].value if len(row) >= 5 else ""
        articularidad = row[5].value if len(row) >= 6 else ""
        nombre = cell.value
        url = cell.hyperlink.target if cell.hyperlink else None
        if nombre and url:
            ejercicios.append({
                "id": len(ejercicios)+1,
                "nombre": nombre,
                "url": url,
                "zona_corporal": zona,
                "implemento": implemento,
                "articularidad": articularidad
            })
    return pd.DataFrame(ejercicios)

df_ejercicios = cargar_datos_excel("All exercices.xlsx")

# --- CONEXION A BASE DE DATOS ---
def obtener_rutina(nombre_clave):
    conn = sqlite3.connect(DB_PATH)
    cur = conn.cursor()
    cur.execute("SELECT id_ejercicio, repeticiones, series FROM rutinas WHERE nombre_clave = ?", (nombre_clave,))
    datos = cur.fetchall()
    conn.close()
    return pd.DataFrame(datos, columns=["id_ejercicio", "repeticiones", "series"])

# --- INICIO DE SESION ---
modo = st.sidebar.radio("Selecciona el modo:", ["Usuario", "Administrador"])

if modo == "Usuario":
    nombre_usuario = st.sidebar.text_input("Ingresa tu nombre_clave")
    if nombre_usuario:
        st.title(f"Rutina personalizada para {nombre_usuario}")
        df_rutina = obtener_rutina(nombre_usuario)
        df_final = df_rutina.merge(df_ejercicios, left_on="id_ejercicio", right_on="id")

        for _, row in df_final.iterrows():
            st.markdown(f"**{row['nombre']}**")
            st.markdown(f"Repeticiones: {row['repeticiones']} | Series: {row['series']}")
            if st.button(f"Ver Video de {row['nombre']}", key=row['id']):
                webbrowser.open_new_tab(row['url'])

elif modo == "Administrador":
    password = st.sidebar.text_input("Contraseña secreta", type="password")
    if password == "dioses123":
        st.title("Panel de Administrador")
        conn = sqlite3.connect(DB_PATH)
        cur = conn.cursor()

        # Crear nuevo usuario
        st.markdown("### ➕ Crear nuevo usuario")
        nuevo_nombre = st.text_input("Nombre_clave nuevo")
        if st.button("Agregar usuario"):
            if nuevo_nombre:
                try:
                    cur.execute("INSERT INTO usuarios (nombre_clave) VALUES (?)", (nuevo_nombre,))
                    conn.commit()
                    st.success(f"Usuario '{nuevo_nombre}' creado.")
                except sqlite3.IntegrityError:
                    st.error("Ese usuario ya existe.")

        # Eliminar usuario existente
        st.markdown("### ❌ Eliminar usuario")
        cur.execute("SELECT nombre_clave FROM usuarios")
        usuarios = [row[0] for row in cur.fetchall()]
        usuario_eliminar = st.selectbox("Seleccionar usuario", usuarios)
        if st.button("Eliminar usuario"):
            cur.execute("DELETE FROM usuarios WHERE nombre_clave = ?", (usuario_eliminar,))
            cur.execute("DELETE FROM rutinas WHERE nombre_clave = ?", (usuario_eliminar,))
            conn.commit()
            st.success(f"Usuario '{usuario_eliminar}' eliminado.")

        
        # Obtener lista actualizada de usuarios
        cur.execute("SELECT nombre_clave FROM usuarios")
        usuarios = [row[0] for row in cur.fetchall()]

        # Mostrar formulario solo si hay usuarios
        if usuarios:
            st.markdown("### ✏️ Asignar ejercicios personalizados")
            usuario_seleccionado = st.selectbox("Selecciona usuario", usuarios, key="mod")

            zona_filtro = st.selectbox("Filtrar por zona corporal", ["Todas"] + sorted(df_ejercicios["zona_corporal"].dropna().unique()))
            implemento_filtro = st.selectbox("Filtrar por implemento", ["Todos"] + sorted(df_ejercicios["implemento"].dropna().unique()))

            df_filtrado = df_ejercicios.copy()
            if zona_filtro != "Todas":
                df_filtrado = df_filtrado[df_filtrado["zona_corporal"] == zona_filtro]
            if implemento_filtro != "Todos":
                df_filtrado = df_filtrado[df_filtrado["implemento"] == implemento_filtro]

            ejercicios_seleccionados = st.multiselect("Selecciona ejercicios", df_filtrado["nombre"].tolist())

            repeticiones = st.number_input("N° Repeticiones", min_value=1, max_value=100, value=10)
            series = st.number_input("N° Series", min_value=1, max_value=20, value=3)

            if st.button("Guardar rutina personalizada"):
                if usuario_seleccionado:
                    cur.execute("DELETE FROM rutinas WHERE nombre_clave = ?", (usuario_seleccionado,))
                    for nombre_ejercicio in ejercicios_seleccionados:
                        id_ej = int(df_ejercicios[df_ejercicios["nombre"] == nombre_ejercicio]["id"].values[0])
                        cur.execute("INSERT INTO rutinas (nombre_clave, id_ejercicio, repeticiones, series) VALUES (?, ?, ?, ?)",
                                    (usuario_seleccionado, id_ej, repeticiones, series))
                    conn.commit()
                    st.success("Rutina guardada exitosamente.")

        conn.close()

    else:
        st.warning("Contraseña incorrecta")
