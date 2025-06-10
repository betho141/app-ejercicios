
import psycopg2
import pandas as pd
from openpyxl import load_workbook

# --- Cargar datos desde Excel ---
def cargar_datos_excel(path):
    wb = load_workbook(path, data_only=True)
    ws = wb.active
    ejercicios = []
    for row in ws.iter_rows(min_row=2):
        cell = row[0]
        zona = row[2].value if len(row) >= 3 else ""
        implemento = row[4].value if len(row) >= 5 else ""
        articularidad = row[5].value if len(row) >= 6 else ""
        nombre = cell.value
        url = cell.hyperlink.target if cell.hyperlink else None
        if nombre and url:
            ejercicios.append((nombre, url, zona, implemento, articularidad))
    return ejercicios

# --- Conexión a Supabase ---
conn = psycopg2.connect(
    host="aws-0-us-east-2.pooler.supabase.com",
    dbname="postgres",
    user="postgres.nwjuqqfsquvytbwslqrw",
    password="Dibujolavida141",
    port=6543
)
cur = conn.cursor()

# --- Cargar y subir datos ---
ejercicios = cargar_datos_excel("All exercices.xlsx")

for ej in ejercicios:
    try:
        cur.execute("""
            INSERT INTO ejercicios (nombre, url, zona_corporal, implemento, articularidad)
            VALUES (%s, %s, %s, %s, %s)
        """, ej)
    except Exception as e:
        print("Error:", ej[0], e)

conn.commit()
cur.close()
conn.close()

print("✅ Datos de ejercicios subidos correctamente a Supabase.")
