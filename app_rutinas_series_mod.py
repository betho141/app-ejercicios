
import streamlit as st
import pandas as pd
import sqlite3
import webbrowser

# --- CONFIGURACIÓN ---
st.set_page_config(page_title="App de Ejercicios", layout="centered")

# --- CARGA DE DATOS ---
def cargar_datos_excel(path):
    from openpyxl import load_workbook
    wb = load_workbook(path, data_only=True)
    ws = wb.active
    ejercicios = []
    for i, row in enumerate(ws.iter_rows(min_row=2), start=1):
        cell = row[0]
        zona = row[2].value if len(row) >= 3 else ""
        implemento = row[4].value if len(row) >= 5 else ""
        articularidad = row[5].value if len(row) >= 6 else ""
        nombre = cell.value
        url = cell.hyperlink.target if cell.hyperlink else None
        if nombre and url:
            ejercicios.append({
                "id": i,
                "nombre": nombre,
                "url": url,
                "zona_corporal": zona,
                "implemento": implemento,
                "articularidad": articularidad
            })
    return pd.DataFrame(ejercicios)

df_ejercicios = cargar_datos_excel("All exercices.xlsx")

# --- INTERFAZ DE LOGIN ---
st.title("🏋️‍♀️ Rutina Personalizada de Ejercicios")
nombre_clave = st.text_input("🔐 Ingrese su nombre_clave")

if nombre_clave:
    conn = sqlite3.connect("rutinas_personalizadas.db")
    cur = conn.cursor()
    cur.execute("SELECT id FROM usuarios WHERE nombre_clave = ?", (nombre_clave,))
    result = cur.fetchone()

    if nombre_clave == "dioses123":
        st.success("Bienvenido Administrador")

        cur.execute("SELECT nombre_clave FROM usuarios")
        usuarios = [u[0] for u in cur.fetchall()]
        usuario_seleccionado = st.selectbox("Selecciona un usuario para asignar rutina", usuarios)

        zona_filtro = st.selectbox("Filtrar por zona corporal", ["Todas"] + sorted(df_ejercicios["zona_corporal"].dropna().unique()))
        implemento_filtro = st.selectbox("Filtrar por implemento", ["Todos"] + sorted(df_ejercicios["implemento"].dropna().unique()))
        articularidad_filtro = st.selectbox("Filtrar por articularidad", ["Todas"] + sorted(df_ejercicios["articularidad"].dropna().unique()))

        df_filtrado = df_ejercicios.copy()
        if zona_filtro != "Todas":
            df_filtrado = df_filtrado[df_filtrado["zona_corporal"] == zona_filtro]
        if implemento_filtro != "Todos":
            df_filtrado = df_filtrado[df_filtrado["implemento"] == implemento_filtro]
        if articularidad_filtro != "Todas":
            df_filtrado = df_filtrado[df_filtrado["articularidad"] == articularidad_filtro]

        ejercicios_asignados = []
        cur.execute("SELECT e.id, e.nombre, r.repeticiones, r.series FROM rutinas r JOIN ejercicios e ON r.id_ejercicio = e.id WHERE r.id_usuario = (SELECT id FROM usuarios WHERE nombre_clave = ?)", (usuario_seleccionado,))
        for row in cur.fetchall():
            ejercicios_asignados.append(row[1])
        ejercicios_seleccionados = st.session_state.get("ejercicios_seleccionados", ejercicios_asignados)

        nuevos_ejercicios = st.multiselect("Selecciona los ejercicios a asignar", df_filtrado["nombre"].tolist(), default=ejercicios_seleccionados)
        st.session_state["ejercicios_seleccionados"] = nuevos_ejercicios

        if nuevos_ejercicios:
            st.markdown("### 📝 Ejercicios seleccionados (haz clic para ver el video o eliminar):")
            ejercicios_para_guardar = []
            for nombre in nuevos_ejercicios:
                url = df_ejercicios[df_ejercicios["nombre"] == nombre]["url"].values[0]
                col1, col2, col3, col4 = st.columns([0.35, 0.25, 0.25, 0.15])
                with col1:
                    st.markdown(f"[{nombre}]({url})", unsafe_allow_html=True)
                with col2:
                    rep = st.number_input(f"Reps: {nombre}", min_value=1, value=st.session_state.get(f"rep_{nombre}", 10), key=f"rep_{nombre}")
                with col3:
                    ser = st.number_input(f"Series: {nombre}", min_value=1, value=st.session_state.get(f"ser_{nombre}", 3), key=f"ser_{nombre}")
                with col4:
                    if st.button("❌", key=f"rm_{nombre}"):
                        st.session_state["ejercicios_seleccionados"].remove(nombre)
                        st.experimental_rerun()
                ejercicios_para_guardar.append((nombre, rep, ser))

            if st.button("📂 Guardar rutina personalizada"):
                cur.execute("SELECT id FROM usuarios WHERE nombre_clave = ?", (usuario_seleccionado,))
                id_usuario = cur.fetchone()[0]
                cur.execute("DELETE FROM rutinas WHERE id_usuario = ?", (id_usuario,))
                for nombre_ejercicio, rep, ser in ejercicios_para_guardar:
                    id_ej = int(df_ejercicios[df_ejercicios["nombre"] == nombre_ejercicio]["id"].values[0])
                    cur.execute("INSERT INTO rutinas (id_usuario, id_ejercicio, repeticiones, series) VALUES (?, ?, ?, ?)",
                                (id_usuario, id_ej, rep, ser))
                conn.commit()
                st.success("Rutina guardada exitosamente.")

    elif result:
        id_usuario = result[0]
        st.success(f"Bienvenido {nombre_clave}")
        cur.execute("""
            SELECT e.nombre, e.url, r.repeticiones, r.series
            FROM rutinas r
            JOIN ejercicios e ON r.id_ejercicio = e.id
            WHERE r.id_usuario = ?
        """, (id_usuario,))
        ejercicios = cur.fetchall()
        st.markdown("### Tu rutina personalizada:")
        for nombre, url, rep, ser in ejercicios:
            st.markdown(f"- [{nombre}]({url}) — **{rep} repeticiones x {ser} series**")
    else:
        st.error("Usuario no encontrado.")
