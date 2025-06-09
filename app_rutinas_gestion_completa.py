
import streamlit as st
import pandas as pd
import sqlite3
import webbrowser
from openpyxl import load_workbook

# --- CONFIGURACIÓN INICIAL ---
st.set_page_config(page_title="App de Ejercicios", layout="centered")

# --- LOGOS ---
def agregar_logo_lateral(imagen_url):
    st.markdown(
        f"""
        <style>
        .stApp {{
            background-image: url("{imagen_url}");
            background-repeat: no-repeat;
            background-attachment: fixed;
            background-size: 300px;
            background-position: 20px 40px;
            padding-left: 340px;
        }}
        </style>
        """,
        unsafe_allow_html=True
    )

def agregar_logo_derecho(imagen_url):
    st.markdown(
        f"""
        <style>
        .stApp::after {{
            content: "";
            background: url("{imagen_url}") no-repeat;
            background-size: 300px;
            position: fixed;
            top: 40px;
            right: 20px;
            width: 300px;
            height: 300px;
            z-index: 1;
        }}
        </style>
        """,
        unsafe_allow_html=True
    )

agregar_logo_lateral("https://i.imgur.com/N2ZpK7M.png")
agregar_logo_derecho("https://i.imgur.com/N2ZpK7M.png")

# --- CONEXIÓN Y FUNCIONES DE BASE DE DATOS ---
def obtener_id_usuario(nombre_clave):
    conn = sqlite3.connect("rutinas_personalizadas.db")
    cur = conn.cursor()
    cur.execute("SELECT id_usuario FROM usuarios WHERE nombre_clave = ?", (nombre_clave,))
    result = cur.fetchone()
    conn.close()
    return result[0] if result else None

def obtener_ids_ejercicios_por_usuario(id_usuario):
    conn = sqlite3.connect("rutinas_personalizadas.db")
    cur = conn.cursor()
    cur.execute("SELECT id_ejercicio FROM rutinas WHERE id_usuario = ?", (id_usuario,))
    resultados = [r[0] for r in cur.fetchall()]
    conn.close()
    return resultados

def asignar_rutina(id_usuario, ids_ejercicios):
    conn = sqlite3.connect("rutinas_personalizadas.db")
    cur = conn.cursor()
    cur.execute("DELETE FROM rutinas WHERE id_usuario = ?", (id_usuario,))
    cur.executemany("INSERT INTO rutinas (id_usuario, id_ejercicio) VALUES (?, ?)",
                    [(id_usuario, ide) for ide in ids_ejercicios])
    conn.commit()
    conn.close()

def agregar_usuario(nombre_clave):
    conn = sqlite3.connect("rutinas_personalizadas.db")
    cur = conn.cursor()
    try:
        cur.execute("INSERT INTO usuarios (nombre_clave) VALUES (?)", (nombre_clave,))
        conn.commit()
        return True, "Usuario agregado correctamente."
    except sqlite3.IntegrityError:
        return False, "Ese nombre_clave ya existe."
    finally:
        conn.close()

def eliminar_usuario(nombre_clave):
    conn = sqlite3.connect("rutinas_personalizadas.db")
    cur = conn.cursor()
    cur.execute("SELECT id_usuario FROM usuarios WHERE nombre_clave = ?", (nombre_clave,))
    usuario = cur.fetchone()
    if usuario:
        id_usuario = usuario[0]
        cur.execute("DELETE FROM rutinas WHERE id_usuario = ?", (id_usuario,))
        cur.execute("DELETE FROM usuarios WHERE id_usuario = ?", (id_usuario,))
        conn.commit()
        result = True
    else:
        result = False
    conn.close()
    return result

# --- CARGA DE EJERCICIOS DESDE EXCEL ---
@st.cache_data
def cargar_ejercicios(path):
    wb = load_workbook(path, data_only=True)
    ws = wb.active
    ejercicios = []
    for idx, row in enumerate(ws.iter_rows(min_row=2), start=1):
        cell = row[0]
        zona = row[2].value if len(row) >= 3 else ""
        implemento = row[4].value if len(row) >= 5 else ""
        articularidad = row[5].value if len(row) >= 6 else ""
        nombre = cell.value
        url = cell.hyperlink.target if cell.hyperlink else None
        if nombre and url:
            ejercicios.append({
                "id_ejercicio": idx,
                "nombre": nombre,
                "url": url,
                "zona_corporal": zona,
                "implemento": implemento,
                "articularidad": articularidad
            })
    return pd.DataFrame(ejercicios)

df_ejercicios = cargar_ejercicios("All exercices.xlsx")

# --- INICIO ---
st.title("🏋️‍♂️ App de Ejercicios Personalizados")

rol = st.selectbox("Selecciona tu rol:", ["Usuario", "Administrador"])

if rol == "Usuario":
    nombre_clave = st.text_input("🔐 Ingresa tu nombre clave:")
    if nombre_clave:
        id_usuario = obtener_id_usuario(nombre_clave)
        if id_usuario:
            st.success(f"Bienvenido, {nombre_clave} 💪")
            ids_permitidos = obtener_ids_ejercicios_por_usuario(id_usuario)
            df = df_ejercicios[df_ejercicios["id_ejercicio"].isin(ids_permitidos)]

            busqueda = st.text_input("🔍 Buscar ejercicio por nombre")
            if busqueda:
                df = df[df["nombre"].str.contains(busqueda, case=False)]

            st.subheader("📊 Tus Ejercicios Asignados")
            st.dataframe(df[["id_ejercicio", "nombre", "zona_corporal", "implemento", "articularidad"]]
                         .rename(columns={
                             "nombre": "Ejercicio",
                             "zona_corporal": "Zona Corporal",
                             "implemento": "Implemento",
                             "articularidad": "Articularidad"
                         }), use_container_width=True)

            st.subheader("🎥 Ver Videos")
            for _, row in df.iterrows():
                if st.button(row["nombre"]):
                    webbrowser.open_new_tab(row["url"])
        else:
            st.error("Nombre clave incorrecto.")

elif rol == "Administrador":
    password = st.text_input("🔐 Contraseña del administrador:", type="password")
    if password == "dioses123":
        st.success("Acceso concedido como administrador 🛠️")

        st.markdown("## ➕ Gestión de Usuarios")
        nuevo_nombre = st.text_input("👤 Ingrese nuevo nombre_clave para agregar")
        if st.button("➕ Agregar usuario"):
            ok, mensaje = agregar_usuario(nuevo_nombre)
            st.success(mensaje) if ok else st.error(mensaje)

        st.markdown("### ❌ Eliminar usuario")
        usuarios_df = pd.read_sql("SELECT * FROM usuarios", sqlite3.connect("rutinas_personalizadas.db"))
        usuario_a_eliminar = st.selectbox("Selecciona un usuario a eliminar:", usuarios_df["nombre_clave"].tolist())
        if st.button("❌ Eliminar usuario seleccionado"):
            if eliminar_usuario(usuario_a_eliminar):
                st.success(f"Usuario '{usuario_a_eliminar}' eliminado correctamente.")
            else:
                st.error("No se pudo eliminar el usuario.")

        st.markdown("## ✍️ Asignar/Modificar Rutina")
        usuarios_df = pd.read_sql("SELECT * FROM usuarios", sqlite3.connect("rutinas_personalizadas.db"))
        selected_user = st.selectbox("Selecciona un usuario para asignar rutina:", usuarios_df["nombre_clave"].tolist())
        id_usuario = usuarios_df[usuarios_df["nombre_clave"] == selected_user]["id_usuario"].values[0]

        zona_opciones = ["Todas"] + sorted(df_ejercicios["zona_corporal"].dropna().unique())
        implemento_opciones = ["Todos"] + sorted(df_ejercicios["implemento"].dropna().unique())

        zona_seleccionada = st.selectbox("🧍 Zona corporal", zona_opciones)
        implemento_seleccionado = st.selectbox("🏋️ Implemento", implemento_opciones)

        df_filtrado = df_ejercicios.copy()
        if zona_seleccionada != "Todas":
            df_filtrado = df_filtrado[df_filtrado["zona_corporal"] == zona_seleccionada]
        if implemento_seleccionado != "Todos":
            df_filtrado = df_filtrado[df_filtrado["implemento"] == implemento_seleccionado]

        st.markdown("### Selecciona los ejercicios que deseas asignar:")
        selected_ejercicios = st.multiselect(
            "Ejercicios disponibles:",
            options=df_filtrado["nombre"].tolist()
        )

        ids_seleccionados = df_ejercicios[df_ejercicios["nombre"].isin(selected_ejercicios)]["id_ejercicio"].tolist()

        if st.button("💾 Guardar rutina personalizada"):
            asignar_rutina(id_usuario, ids_seleccionados)
            st.success("Rutina actualizada correctamente.")
    elif password:
        st.error("Contraseña incorrecta.")
